import math
import openpyxl

from collections import defaultdict
from datetime import datetime, date, time
from io import BytesIO
from openpyxl.utils import get_column_letter

from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render, redirect

from ..constants import constants_view
from ..models import (
    RMOrder, RMCustomer, OrderImage, Container, RMProduct, OrderItem,
    AlineOrderRecord, ContainerItem, UserAndPermission, Employee
)

from .utils.getPermission import get_user_permissions
from .utils.date_utils import sort_by_date

def inventory_view(request):
    employee_id = request.GET.get('employee')    
    inventory_items = RMProduct.objects.filter(type = "Rimei")
    if employee_id:
        inventory_items = inventory_items.filter(blongTo_id=employee_id)
    
    inventory_items_converty = []  
    for product in inventory_items:
        # 查询库存记录
        inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list = get_quality(product)
        productTemp = get_product_qty(product, inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list)

        inventory_items_converty.append(productTemp)
    inventory_items_converty = sorted(inventory_items_converty, key=lambda x: x.name)

    employees = Employee.objects.filter(belongTo = "CabinetsDepot")
    user_permissions = get_user_permissions(request.user)
    return render(request, constants_view.template_inventory, {
        "inventory_items": inventory_items_converty,
        'employees': employees,
        'user_permissions': user_permissions,
    })
    
def inventory_diff_view(request):
    inventory_items = RMProduct.objects.filter(type = "Rimei")

    diff_items = []  # 用来存储有差异的库存记录
    for product in inventory_items:
        # 查询库存记录
        inbound_list, outbound_list, outbound_actual_list,outbound_stock_list, inbound_actual_list = get_quality(product)
        productTemp = get_product_qty(product, inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list)

        if productTemp.quantity != productTemp.quantity_for_neworder:
            diff_items.append(productTemp)

    diff_items = sorted(diff_items, key=lambda x: x.quantity_for_neworder)

    user_permissions = get_user_permissions(request.user)
    return render(request, constants_view.template_inventory, {"inventory_items": diff_items,'user_permissions': user_permissions})

def inventory_metal(request):
    employee_id = request.GET.get('employee')
    inventory_items = RMProduct.objects.filter(type = "Metal")
    
    if employee_id:
        inventory_items = inventory_items.filter(blongTo_id=employee_id)

    inventory_items_converty = []  # 用来存储有差异的库存记录
    for product in inventory_items:
        # 查询库存记录
        inbound_list, outbound_list, outbound_actual_list,outbound_stock_list, inbound_actual_list = get_quality(product)

        productTemp = get_product_qty(product, inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list)

        inventory_items_converty.append(productTemp)
    inventory_items_converty = sorted(inventory_items_converty, key=lambda x: x.name)

    employees = Employee.objects.filter(belongTo = "CabinetsDepot")
    user_permissions = get_user_permissions(request.user)
    return render(request, constants_view.template_inventory, {
        "inventory_items": inventory_items_converty,
        'employees': employees,
        'user_permissions': user_permissions})

def inventory_diff_metal(request):
    employee_id = request.GET.get('employee')
    inventory_items = RMProduct.objects.filter(type = "Metal")
    
    if employee_id:
        inventory_items = inventory_items.filter(blongTo_id=employee_id)

    inventory_items_converty = []  # 用来存储有差异的库存记录
    for product in inventory_items:
        # 查询库存记录
        inbound_list, outbound_list, outbound_actual_list,outbound_stock_list, inbound_actual_list = get_quality(product)

        productTemp = get_product_qty(product, inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list)
        # if productTemp.quantity != productTemp.quantity_for_neworder:
        if productTemp.status == 'Short':
            inventory_items_converty.append(productTemp)
    inventory_items_converty = sorted(inventory_items_converty, key=lambda x: x.name)

    employees = Employee.objects.filter(belongTo = "CabinetsDepot")
    user_permissions = get_user_permissions(request.user)
    return render(request, constants_view.template_inventory, {
        "inventory_items": inventory_items_converty,
        'employees': employees,
        'user_permissions': user_permissions})


def inventory_mcd(request):
    inventory_items = RMProduct.objects.filter(Q(type="Mcdonalds") | Q(type="Other") | Q(type="MCD"))

    inventory_items_converty = []  # 用来存储有差异的库存记录
    for product in inventory_items:
        # 查询库存记录
        inbound_list, outbound_list, outbound_actual_list,outbound_stock_list, inbound_actual_list = get_quality(product)

        productTemp = get_product_qty(product, inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list)

        inventory_items_converty.append(productTemp)
        inventory_items_converty = sorted(inventory_items_converty, key=lambda x: (x.type,x.name))

    user_permissions = get_user_permissions(request.user)
    return render(request, constants_view.template_inventory, {"inventory_items": inventory_items_converty,'user_permissions': user_permissions})

def export_stock(request):
    inventory_items = RMProduct.objects.all()
    print("----------export_stock------------", len(inventory_items))
    inventory_items_converty = []

    for product in inventory_items:
        inbound_list, outbound_list, outbound_actual_list, outbound_stock_list, inbound_actual_list = get_quality(product)

        productTemp = get_product_qty(product, inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list)

        inventory_items_converty.append(productTemp)

    inventory_items_converty = sorted(inventory_items_converty, key=lambda x: (x.type, x.name))
    user_permissions = get_user_permissions(request.user)

    # 🔽 Export to Excel if requested
    if request.GET.get("export") == "1":
        return export_inventory_to_excel(inventory_items_converty)

    return render(request, constants_view.template_inventory, {
        "inventory_items": inventory_items_converty,
        "user_permissions": user_permissions
    })

def export_metal_stock(request):
    inventory_items = RMProduct.objects.all()
    print("----------export_metal_stock------------", len(inventory_items))
    inventory_items_converty = []

    for product in inventory_items:
        inbound_list, outbound_list, outbound_actual_list, outbound_stock_list, inbound_actual_list = get_quality(product)

        productTemp = get_product_qty(product, inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list)

        inventory_items_converty.append(productTemp)

    inventory_items_converty = sorted(inventory_items_converty, key=lambda x: (x.type, x.name))
    user_permissions = get_user_permissions(request.user)

    # 🔽 Export to Excel if requested
    if request.GET.get("export") == "1":
        return export_metal_inventory_to_excel(inventory_items_converty)

    return render(request, constants_view.template_inventory, {
        "inventory_items": inventory_items_converty,
        "user_permissions": user_permissions
    })
    

def order_history(request,product_id):
    product = get_object_or_404(RMProduct, id=product_id)

    # 查询库存记录
    inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list = get_quality(product)

    # 计算入库总数量
    total_inbound_quantity = sum(item['quantity'] or 0 for item in inbound_list)
    # 计算出库总数量
    total_outbound_quantity = sum(item['quantity'] or 0 for item in outbound_list)
    # 计算实际入库数量
    total_inbound_actual_quantity = sum(item['quantity'] or 0 for item in inbound_actual_list)
    # 计算实际出库数量
    total_outbound_actual_quantity = sum(item['quantity'] or 0 for item in outbound_actual_list)
    # 计算备货区数量
    total_outbound_stock_quantity = sum(item['quantity'] or 0 for item in outbound_stock_list)

    total_surplus_quantity = total_inbound_quantity - total_outbound_quantity
    total_quality = total_inbound_actual_quantity - total_outbound_actual_quantity
    total_stock = total_inbound_actual_quantity - total_outbound_actual_quantity - total_outbound_stock_quantity

    return render(request, constants_view.template_order_history, {
        'product': product,
        'inbound_logs': inbound_list,
        'outbound_logs': outbound_list,
        'total_inbound_quantity': total_inbound_quantity,
        'total_outbound_quantity': total_outbound_quantity,
        'total_surplus_quantity': total_surplus_quantity,
        'total_inbound_actual_quantity':total_inbound_actual_quantity,
        'total_outbound_actual_quantity':total_outbound_actual_quantity,
        'total_quantity': total_quality,
        'total_stock': total_stock,
        "today": date.today(),
    })

def inventory_summary(request):
    
    # print("----------inventory_summary------------",len(inventory_items))
    inventory_items = RMProduct.objects.all()  # 获取所有库存信息
    summary_by_type = {} 
    
    for product in inventory_items:
        product_type = product.type or "Unknown"

        # 查询库存记录
        inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list = get_quality(product)
        productTemp = get_product_qty(product, inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list)
        # total_pallets += math.ceil(product.quantity / product.Pallet) # 数量，向上取整

        if product_type not in summary_by_type:
            summary_by_type[product_type] = {
                'employee_data': {},
                'total_pallets': 0,
            }

        # ===== 计算托盘数（向上取整）=====
        if product.Pallet:
            summary_by_type[product_type]['total_pallets'] += math.ceil(
                product.quantity / product.Pallet
            )

        # ===== 原有逻辑 =====
        inbound_list, outbound_list, outbound_actual_list, outbound_stock_list, inbound_actual_list = get_quality(product)
        productTemp = get_product_qty(
            product,
            inbound_list,
            outbound_list,
            outbound_actual_list,
            outbound_stock_list,
            inbound_actual_list
        )

        employee = product.blongTo.name if product.blongTo else "Unknown"
        qty = product.quantity

        employee_data = summary_by_type[product_type]['employee_data']

        if employee not in employee_data:
            employee_data[employee] = {
                'total_qty': 0,
                'qty_list': [],
                'product_set': set(),
                'total_diff': 0,
            }

        employee_data[employee]['total_qty'] += qty
        employee_data[employee]['total_diff'] += product.quantity_diff
        employee_data[employee]['qty_list'].append(qty)
        employee_data[employee]['product_set'].add(product.id)

    # ===== 构造每个 type 的 summary =====
    final_summary = {}

    for product_type, data in summary_by_type.items():
        employee_data = data['employee_data']

        total_qty_all = sum(v['total_qty'] for v in employee_data.values())
        total_qty_diff = sum(v['total_diff'] for v in employee_data.values())

        summary = []
        for employee, v in employee_data.items():
            percentage = round((v['total_qty'] / total_qty_all) * 100, 2) if total_qty_all else 0
            summary.append({
                "employee": employee,
                "total_qty": v['total_qty'],
                "total_diff": v['total_diff'],
                "qty_expression": ' + '.join(str(q) for q in v['qty_list']),
                "product_count": len(v['product_set']),
                "percentage": percentage,
            })

        summary = sorted(summary, key=lambda x: x['percentage'], reverse=True)

        final_summary[product_type] = {
            "summary": summary,
            "total_qty_all": total_qty_all,
            "total_qty_diff": total_qty_diff,
            "total_pallets": data['total_pallets'],
        }
    
    user_permissions = get_user_permissions(request.user)
    years = [date.today().year]
    months = list(range(1, 13))  # 1 到 12 月
    return render(request, constants_view.template_temporary, {
        'user_permissions': user_permissions,
        'years':years,'months':months,
        'final_summary': final_summary, 
    })



def edit_product(request,product_id):
    product = get_object_or_404(RMProduct, id=product_id)
    employee = Employee.objects.filter(belongTo = "CabinetsDepot")

    if request.method == 'POST':
        product.name = request.POST.get('name')
        product.shortname = request.POST.get('shortname')
        product.size = request.POST.get('size')
        product.TI = request.POST.get('TI') or 0
        product.HI = request.POST.get('HI') or 0
        product.Pallet = request.POST.get('Pallet') or 0
        product.Color = request.POST.get('Color')
        product.Location = request.POST.get('Location')
        product.ShelfRecord = request.POST.get('ShelfRecord')
        product.description = request.POST.get('description')
        product.quantity_init = request.POST.get('quantity_init') or 0
        product.quantity_diff = request.POST.get('quantity_diff') or 0
        product.type = request.POST.get('type')
        product.price = request.POST.get('price') or 0
        product.blongTo = Employee.objects.get(id=request.POST.get('blongTo'))
        product.status = request.POST.get('status')

        product.save()

        if product.type == 'Metal':
            return redirect('inventory_metal')
        elif product.type == 'Mcdonalds' or product.type == 'MCD':
            return redirect('inventory_mcd')
        else:
            return redirect('inventory')

    return render(request, constants_view.template_edit_product, {
        'product': product,
        'employee': employee,
    })

# sub function
def get_month_pallet_number(select_month, select_year):
        # 过滤出指定月份的订单
    start_date = datetime(int(select_year), int(select_month), 1)
    if select_month == '12':
        end_date = datetime(int(select_year) + 1, 1, 1)  # 下一年1月1日
    else:
        end_date = datetime(int(select_year), int(select_month) + 1, 1)  # 下个月的1日

    # 排除 	GoldenFeather:3, Metal Studs:12, Other Freight Forwarders / Omar Transfer:6
    container_in_orders = Container.objects.filter(
        empty_date__gte=start_date, 
        empty_date__lt=end_date
    # ).exclude(Q(customer="6")).order_by('-customer')
    ).exclude(Q(customer="3") | Q(customer="6")| Q(customer="12")).order_by('empty_date')
    print("container in : ",len(container_in_orders))

    # 查询 "gloves in" 数据
    # 排除 	GoldenFeather：3,	Metal Studs：12
    gloves_in_orders = Container.objects.filter(
        empty_date__gte=start_date, 
        empty_date__lt=end_date
    ).exclude(Q(customer="3") | Q(customer="12")).order_by('-customer')
    print("golve in : ",len(gloves_in_orders))

    # 创建 "gloves in" DataFrame
    gloves_in_data = [{
        'inbound_date': order.empty_date,
        'container_id': order.container_id,
        'plts': order.plts,
        'customer': order.customer,
        'description': order.content
        }for order in gloves_in_orders
    ]

    # 查询 "gloves out" 数据（根据您的需求进行调整）
    # 排除 Omar Sample:8, 	OBL:19
    # gloves_out_orders = RMOrder.objects.filter(
    #     outbound_date__gte=start_date, 
    #     outbound_date__lt=end_date
    # ).exclude(Q(customer_name="8") | Q(customer_name="19")).order_by('outbound_date')
    gloves_out_orders = RMOrder.objects.filter(
        pickup_date__gte=start_date, 
        pickup_date__lt=end_date
    ).exclude(Q(customer_name="8") | Q(customer_name="19")).order_by('pickup_date')
    print("golve out : ",len(gloves_out_orders))

    # ✅ 计算 Container in 总数
    total_container = len(container_in_orders)

    # ✅ 计算 gloves in 总托盘数
    total_in_plts = sum(order.plts for order in gloves_in_orders)

    # ✅ 计算 gloves out 总托盘数
    total_out_plts = sum(order.plts for order in gloves_out_orders)

    return total_container,total_in_plts,total_out_plts, gloves_in_data, container_in_orders
    
def get_product_qty(product, inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list):
     
      # 计算入库总数量
    total_inbound_quantity = sum(item['quantity'] or 0 for item in inbound_list)
    # 计算出库总数量
    total_outbound_quantity = sum(item['quantity'] for item in outbound_list)
    # 计算实际入库数量
    total_inbound_actual_quantity = sum(item['quantity'] or 0 for item in inbound_actual_list)
    # 计算实际出库数量
    total_outbound_actual_quantity = sum(item['quantity'] for item in outbound_actual_list)
    # 计算备货区数量
    total_outbound_stock_quantity = sum(item['quantity'] for item in outbound_stock_list)

    product.quantity_for_neworder = total_inbound_actual_quantity - total_outbound_quantity
    product.quantity = total_inbound_actual_quantity - total_outbound_actual_quantity
    product.quantity_to_stock = product.quantity - total_outbound_stock_quantity
    product.quantity_outbound_stock = total_outbound_stock_quantity
    product.shownumber  = product.quantity_to_stock + product.quantity_diff
    product.availableQTY = product.quantity_for_neworder + product.quantity_diff
    product.pallet = product.Pallet
    product.color = product.Color
    product.palletnumber = 0 if product.Pallet == 0 else product.shownumber // product.Pallet
    product.case = 0 if product.Pallet == 0 else product.shownumber % product.Pallet
    product.Location = product.Location
    product.ShelfRecord = product.ShelfRecord or '-'
    # product.totalPallet = 0 if product.Pallet == 0 else math.ceil(product.quantity / product.Pallet)

    return product

def get_quality(product):
    
    # 创建一个“初始记录”对象，模拟一个类似入库日志的结构
    initial_log = {
        'date': 'Initial',
        'quantity': product.quantity_init,
        'operator': 'System',
        'note': 'Initial stock quantity'
    }

    # 入库记录：ContainerItem + Container 的 delivery_date
    inbound_logs = ContainerItem.objects.filter(product=product).select_related('container')
    inbound_list = [{
        'date': item.container.delivery_date if item.container else None,
        'quantity': item.quantity,
        'operator': getattr(item.container, 'created_user', str(item.container)) if hasattr(item, 'container') else 'N/A',
        'note': getattr(item.container, 'container_id', '')
    } for item in inbound_logs]

    # 排序
    inbound_list = sorted(inbound_list, key=lambda x: sort_by_date(x, "date"), reverse=True)

    # 插入初始记录到入库列表最前
    inbound_list.insert(0, initial_log)    

    # 所有出库记录：OrderItem + RMOrder 的 outbound_date
    outbound_logs = OrderItem.objects.filter(
        product=product,
        order__is_canceled=False  # 排除已取消订单
    ).select_related('order', 'order__customer_name')
    outbound_list = [{
        'order_date': item.order.order_date if item.order and item.order.order_date else 'N/A',
        'pickup_date': item.order.pickup_date if item.order and item.order.pickup_date else 'N/A',
        'quantity': item.quantity,
        'operator': getattr(item.order, 'created_user', 'N/A'),  # 使用 RMOrder 中的 created_user 或者其他字段作为操作员
        'customer_name': getattr(item.order.customer_name, 'name', 'N/A'),  # 正确获取客户名
        'note': getattr(item.order, 'so_num', '')
    } for item in outbound_logs]

    # 排序
    outbound_list = sorted(outbound_list, key=lambda x: sort_by_date(x, "pickup_date"), reverse=True)

    # 实际入库记录
    inbound_actual_logs = ContainerItem.objects.filter(product=product,container__is_updateInventory=True).select_related('container')
    inbound_actual_list = [{
        'date': item.container.delivery_date if item.container and item.container.delivery_date else 'N/A',
        'quantity': item.quantity,
        'operator': getattr(item.container, 'created_user', str(item.container)) if hasattr(item, 'container') else 'N/A',  # 使用 RMOrder 中的 created_user 或者其他字段作为操作员
        'note': getattr(item.container, 'container_id', '')
    } for item in inbound_actual_logs]
    # 插入初始记录到入库列表最前
    inbound_actual_list.insert(0, initial_log)    

    # 实际出库记录
    outbound_actual_logs = OrderItem.objects.filter(
        product=product,
        order__is_updateInventory=True
    ).select_related(
        'order', 'order__customer_name'  # 预加载 order 和 RMCustomer
    )
    outbound_actual_list = [{
        'date': item.order.pickup_date if item.order and item.order.pickup_date else 'N/A',
        'date_shipped': item.order.outbound_date if item.order and item.order.outbound_date else 'N/A',
        'quantity': item.quantity,
        'operator': getattr(item.order, 'created_user', 'N/A'),
        'customer_name': getattr(item.order.customer_name, 'name', 'N/A'),  # 正确获取客户名
        'note': getattr(item.order, 'so_num', '')
    } for item in outbound_actual_logs]

    # 备货区数量
    outbound_stock_logs = OrderItem.objects.filter(product=product,order__is_allocated_to_stock=True).select_related('order')
    outbound_stock_list = [{
        'date': item.order.pickup_date if item.order and item.order.pickup_date else 'N/A',
        'date_shipped': item.order.outbound_date if item.order and item.order.outbound_date else 'N/A',
        'quantity': item.quantity,
        'operator': getattr(item.order, 'created_user', 'N/A'),  # 使用 RMOrder 中的 created_user 或者其他字段作为操作员
        'customer_name': getattr(item.order.customer_name, 'name', 'N/A'),  # 正确获取客户名
        'note': getattr(item.order, 'so_num', '')
    } for item in outbound_stock_logs]

    return inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list

def export_inventory_to_excel(items):
    wb = openpyxl.Workbook()
    ws = wb.active
    today_str = datetime.now().strftime("%Y_%m_%d")
    ws.title = f"Inventory_{today_str}"

    # 按 type 分类
    type_groups = defaultdict(list)
    for item in items:
        type_groups[item.type].append(item)

    # Define headers
    headers = ["Product", "QTY", "QTY removing order", "Diff","Show QTY","Available QTY", "Each","Color", "Location", "Pallets", "Cases","ShelfRecord"]
    # ws.append(headers)

    # Write data rows
    for i, (type_name, grouped_items) in enumerate(type_groups.items()):

        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = str(type_name)[:31] or "Sheet"  # Excel sheet name max 31 chars

        ws.append(headers)

        for item in grouped_items:
            ws.append([
                str(item),
                item.quantity,
                item.quantity_for_neworder,
                item.quantity_diff,
                item.shownumber,
                item.availableQTY,
                item.pallet,
                item.color,
                item.Location,
                item.palletnumber,
                item.case,
                item.ShelfRecord,
            ])

        # 自动列宽
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # Save to in-memory file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create HTTP response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Inventory_{today_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def export_metal_inventory_to_excel(items):
    today_str = datetime.now().strftime("%Y_%m_%d")

    wb = openpyxl.Workbook()
    ws = wb.active    
    ws.title = f"Metal_Inventory_{today_str}"

    # ⭐ 只保留 Metal
    metal_items = [item for item in items if item.type == "Metal"]

    # Define headers
    headers = ["Product", "QTY", "QTY removing order", "ShelfRecord"]
    ws.append(headers)

    for item in metal_items:
        ws.append([
            str(item),
            item.quantity,
            item.quantity_for_neworder,
            item.ShelfRecord,
        ])

    # 自动列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # Save to in-memory file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create HTTP response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Metal_Inventory_{today_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



