from django.shortcuts import render, redirect
from django.contrib import messages
from ..models import RMOrder, RMCustomer, OrderImage, Container, RMProduct,RMInventory, OrderItem, AlineOrderRecord, ContainerItem, UserAndPermission
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from django.http import JsonResponse
from django.db.models import Count
from datetime import datetime
from django.http import HttpResponse
import pandas as pd
import json
from django.utils import timezone
from django.shortcuts import get_object_or_404
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO
from datetime import datetime, date, time

def add_order(request):
    if request.method == "POST":
        try:
            # 检查SO号是否已存在
            so_num = request.POST.get('so_num')
            if RMOrder.objects.filter(so_num=so_num).exists():
                messages.error(request, f'创建订单失败：SO号 {so_num} 已存在')

                customers = RMCustomer.objects.all()
                return render(request, 'container/rmorder/add_order.html', {
                    'so_no': so_num,
                    'po_no': request.POST.get('po_num'),
                    'plts': request.POST.get('plts'),
                    'bill_to': request.POST.get('bill_to'),
                    'ship_to': request.POST.get('ship_to'),
                    "order_date": request.POST.get('order_date'),
                    'pickup_date': request.POST.get('pickup_date'),
                    'outbound_date': request.POST.get('outbound_date'),
                    'customers':customers,
                    'customer_name': RMCustomer.objects.get(id=request.POST.get('customer_name')),
                    'is_sendemail': request.POST.get('is_sendemail') == 'on',
                    'is_updateInventory': request.POST.get('is_updateInventory') == 'on',
                    'is_canceled': request.POST.get('is_canceled') == 'on',
                    'is_allocated_to_stock': request.POST.get('is_allocated_to_stock') == 'on',
                    'order_pdfname':request.POST.get('order_pdfname')  # 添加这一行
                })
            
            print("------add_order-----")
            customer = RMCustomer.objects.get(id=request.POST.get('customer_name'))
            order = RMOrder(
                so_num=request.POST.get('so_num'),
                po_num=request.POST.get('po_num'),
                plts=request.POST.get('plts'),
                customer_name=customer,
                bill_to=request.POST.get('bill_to'),
                ship_to=request.POST.get('ship_to'),
                order_pdfname = request.POST.get('order_pdfname'),
                order_date = request.POST.get('order_date') or None,
                pickup_date=request.POST.get('pickup_date') or None,
                outbound_date=request.POST.get('outbound_date') or None,
                is_sendemail=request.POST.get('is_sendemail') == 'on',
                is_updateInventory=request.POST.get('is_updateInventory') == 'on',
                is_canceled=request.POST.get('is_canceled') == 'on',
                is_allocated_to_stock=request.POST.get('is_allocated_to_stock') == 'on',  
                created_user=request.user,  # ✅ 保存创建人           
            )
            order.save()

            # 假设您从 PDF 中提取的产品信息存储在一个字典中
            order_items_json = request.POST.get('orderitems')  # 获取产品信息
            print("hello??",order_items_json)
            if order_items_json:
                print("hello??")
                order_items = json.loads(order_items_json)  # 解析JSON
                for item in order_items:
                    product_name = item['item']
                    quantity = item['qty']                    
                    print(f"处理商品: {product_name}, 数量: {quantity}")

                    products = RMProduct.objects.all()
                    product = None
                    for p in products:
                        if (p.shortname and p.shortname.strip() in product_name) or (p.name in product_name):
                            product = p
                            break
                    if product:
                        # 创建订单明细
                        OrderItem.objects.create(order=order, product=product, quantity=int(quantity))
                    else:
                        print(f"警告: 未找到匹配的产品 '{product_name}'")
            
            # 更新库存for neworder


            return redirect('rimeiorder')
        except Exception as e:
            messages.error(request, f'创建订单失败：{str(e)}')
    # GET请求处理
    customers = RMCustomer.objects.all()
    return render(request, 'container/rmorder/add_order.html', {'customers': customers})

@require_http_methods(["GET", "POST"])
def edit_order(request, so_num):
    print("--------edit_order------",so_num)
    try:
        if request.method == "GET":
            order = RMOrder.objects.get(so_num=so_num)
            customers = RMCustomer.objects.all()
            order_items = OrderItem.objects.filter(order=order)
            print("---:",len(order_items),order.so_num)
            products = RMProduct.objects.all()
            return render(request, 'container/rmorder/edit_order.html', {
                'order': order,
                'customers': customers,
                'order_items': order_items,
                'products': products,  # 加上这行
            })
        elif request.method == "POST":
            try:
                order = RMOrder.objects.get(so_num=so_num)
                new_so_num = request.POST.get('so_num')
                if new_so_num != so_num and RMOrder.objects.filter(so_num=new_so_num).exists():
                    messages.error(request, f'更新订单失败：SO号 {new_so_num} 已存在')
                    customers = RMCustomer.objects.all()
                    return render(request, 'container/rmorder/edit_order.html', {
                        'order': order,
                        'customers': customers
                    })

                customer = RMCustomer.objects.get(id=request.POST.get('customer_name'))
                order.so_num = new_so_num
                order.po_num = request.POST.get('po_num')
                order.plts = request.POST.get('plts')
                order.customer_name = customer
                order.bill_to = request.POST.get('bill_to')
                order.ship_to = request.POST.get('ship_to')
                order.pickup_date = request.POST.get('pickup_date') or None
                order.outbound_date = request.POST.get('outbound_date') or None
                order.is_sendemail = request.POST.get('is_sendemail') == 'on'
                order.is_updateInventory = request.POST.get('is_updateInventory') == 'on'
                order.is_canceled = request.POST.get('is_canceled') == 'on'
                order.is_allocated_to_stock = request.POST.get('is_allocated_to_stock') == 'on'                                         
                order.save()

                # 更新订单项目
                items_json  = request.POST.get('orderitems')
                print("---:",items_json)
                if items_json :
                    items = json.loads(items_json )

                    # ⚠️ 先清空旧的条目（如果你是编辑页面）
                    OrderItem.objects.filter(order=order).delete()
                    for item in items:
                        product = RMProduct.objects.get(id=item['product_id'])
                        quantity = int(item['quantity'])
                        print("---:",product,quantity)
                        
                        OrderItem.objects.create(
                            order=order,
                            product=product,
                            quantity=int(quantity)
                        )                            

                messages.success(request, '订单更新成功！')
                return redirect('rimeiorder')
            except Exception as e:
                messages.error(request, f'更新订单失败：{str(e)}')
                customers = RMCustomer.objects.all()
                order_items = OrderItem.objects.filter(order=order)
                products = RMProduct.objects.all()
                return render(request, 'container/rmorder/edit_order.html', {
                    'order': order,
                    'customers': customers,
                    'order_items': order_items,
                    'products': products,
                })
        
        
    except RMOrder.DoesNotExist:
        messages.error(request, '订单不存在')
        return redirect('rimeiorder')

def inventory_view(request):
    inventory_items = RMInventory.objects.all()  # 获取所有库存信息
    print("----------inventory_view------------",len(inventory_items))
    inventory_items_converty = []
    for product in inventory_items:
        # 查询库存记录
        inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list = get_quality(product.product)

        product = get_product_qty(product, inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list)

        inventory_items_converty.append(product)

        inventory_items_converty = sorted(inventory_items_converty, key=lambda x: x.product.name)

    user_permissions = get_user_permissions(request.user)
    return render(request, "container/inventory.html", {"inventory_items": inventory_items_converty,'user_permissions': user_permissions})

def inventory_diff_view(request):
    inventory_items = RMInventory.objects.all()  # 获取所有库存信息

    diff_items = []  # 用来存储有差异的库存记录
    for product in inventory_items:
        # 查询库存记录
        inbound_list, outbound_list, outbound_actual_list,outbound_stock_list, inbound_actual_list = get_quality(product.product)

        product = get_product_qty(product, inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list)

        if product.quantity != product.quantity_for_neworder:
            diff_items.append(product)

        diff_items = sorted(diff_items, key=lambda x: x.quantity_for_neworder)

    user_permissions = get_user_permissions(request.user)
    return render(request, "container/inventory.html", {"inventory_items": diff_items,'user_permissions': user_permissions})

def export_stock(request):
    inventory_items = RMInventory.objects.all()
    print("----------export_stock------------", len(inventory_items))
    inventory_items_converty = []

    for product in inventory_items:
        inbound_list, outbound_list, outbound_actual_list, outbound_stock_list, inbound_actual_list = get_quality(product.product)

        product = get_product_qty(product, inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list)

        inventory_items_converty.append(product)

    inventory_items_converty = sorted(inventory_items_converty, key=lambda x: x.product.name)
    user_permissions = get_user_permissions(request.user)

    # 🔽 Export to Excel if requested
    if request.GET.get("export") == "1":
        return export_inventory_to_excel(inventory_items_converty)

    return render(request, "container/inventory.html", {
        "inventory_items": inventory_items_converty,
        "user_permissions": user_permissions
    })

def order_history(request,product_id):
    product = get_object_or_404(RMProduct, id=product_id)

    # 查询库存记录
    inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list = get_quality(product)

    # 计算入库总数量
    total_inbound_quantity = sum(item['quantity'] for item in inbound_list)
    # 计算出库总数量
    total_outbound_quantity = sum(item['quantity'] for item in outbound_list)
    # 计算实际入库数量
    total_inbound_actual_quantity = sum(item['quantity'] for item in inbound_actual_list)
    # 计算实际出库数量
    total_outbound_actual_quantity = sum(item['quantity'] for item in outbound_actual_list)
    # 计算备货区数量
    total_outbound_stock_quantity = sum(item['quantity'] for item in outbound_stock_list)

    total_surplus_quantity = total_inbound_quantity - total_outbound_quantity
    total_quality = total_inbound_actual_quantity - total_outbound_actual_quantity
    total_stock = total_inbound_actual_quantity - total_outbound_actual_quantity - total_outbound_stock_quantity

    return render(request, 'container/inventory/order_history.html', {
        'product': product,
        'inbound_logs': inbound_list,
        'outbound_logs': outbound_list,
        'total_inbound_quantity': total_inbound_quantity,
        'total_outbound_quantity': total_outbound_quantity,
        'total_surplus_quantity': total_surplus_quantity,
        'total_inbound_actual_quantity':total_inbound_actual_quantity,
        'total_outbound_actual_quantity':total_outbound_actual_quantity,
        'total_quantity': total_quality,
        'total_stock': total_stock
    })

def get_product_qty(product, inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list):
     
      # 计算入库总数量
    total_inbound_quantity = sum(item['quantity'] for item in inbound_list)
    # 计算出库总数量
    total_outbound_quantity = sum(item['quantity'] for item in outbound_list)
    # 计算实际入库数量
    total_inbound_actual_quantity = sum(item['quantity'] for item in inbound_actual_list)
    # 计算实际出库数量
    total_outbound_actual_quantity = sum(item['quantity'] for item in outbound_actual_list)
    # 计算备货区数量
    total_outbound_stock_quantity = sum(item['quantity'] for item in outbound_stock_list)

    product.quantity_for_neworder = total_inbound_actual_quantity - total_outbound_quantity
    product.quantity = total_inbound_actual_quantity - total_outbound_actual_quantity
    product.quantity_to_stock = product.quantity - total_outbound_stock_quantity
    product.shownumber  = product.quantity_to_stock + product.quantity_diff
    product.pallet = product.product.Pallet
    product.color = product.product.Color
    product.palletnumber = product.shownumber//product.product.Pallet
    product.case = product.shownumber % product.product.Pallet
    product.Location = product.product.Location
    product.ShelfRecord = product.product.ShelfRecord

    return product

def get_quality(product):
    # 获取 RMInventory 中的初始库存记录
    inventory = RMInventory.objects.filter(product=product).first()
    
    # 创建一个“初始记录”对象，模拟一个类似入库日志的结构
    initial_log = {
        'date': 'Initial',
        'quantity': inventory.quantity_init if inventory else 0,
        'operator': 'System',
        'note': 'Initial stock quantity'
    }

    # 入库记录：ContainerItem + Container 的 empty_date
    inbound_logs = ContainerItem.objects.filter(product=product).select_related('container')
    inbound_list = [{
        'date': item.container.empty_date if item.container else None,
        'quantity': item.quantity,
        'operator': getattr(item.container, 'created_user', str(item.container)) if hasattr(item, 'container') else 'N/A',
        'note': getattr(item.container, 'container_id', '')
    } for item in inbound_logs]

    # 排序
    inbound_list = sorted(inbound_list, key=lambda x: sort_by_date(x, "date"), reverse=True)

    # 插入初始记录到入库列表最前
    inbound_list.insert(0, initial_log)    

    # 出库记录：OrderItem + RMOrder 的 outbound_date
    outbound_logs = OrderItem.objects.filter(
        product=product,
        order__is_canceled=False  # 排除已取消订单
    ).select_related('order')
    outbound_list = [{
        'date': item.order.pickup_date if item.order and item.order.pickup_date else 'N/A',
        'pickup_date': item.order.pickup_date if item.order and item.order.pickup_date else 'N/A',
        'quantity': item.quantity,
        'operator': getattr(item.order, 'created_user', 'N/A'),  # 使用 RMOrder 中的 created_user 或者其他字段作为操作员
        'note': getattr(item.order, 'so_num', '')
    } for item in outbound_logs]

    # 排序
    outbound_list = sorted(outbound_list, key=lambda x: sort_by_date(x, "pickup_date"), reverse=True)

    # 实际入库记录
    inbound_actual_logs = ContainerItem.objects.filter(product=product,container__is_updateInventory=True).select_related('container')
    inbound_actual_list = [{
        'date': item.container.delivery_date if item.container and item.container.delivery_date else 'N/A',
        'quantity': item.quantity,
        'operator': getattr(item.container, 'created_user', str(item.container)) if hasattr(item, 'container') else 'N/A',  # 使用 RMOrder 中的 created_user 或者其他字段作为操作员
        'note': getattr(item.container, 'container_id', '')
    } for item in inbound_actual_logs]
    # 插入初始记录到入库列表最前
    inbound_actual_list.insert(0, initial_log)    

    outbound_actual_logs = OrderItem.objects.filter(product=product,order__is_updateInventory=True).select_related('order')
    outbound_actual_list = [{
        'date': item.order.pickup_date if item.order and item.order.pickup_date else 'N/A',
        'date_shipped': item.order.outbound_date if item.order and item.order.outbound_date else 'N/A',
        'quantity': item.quantity,
        'operator': getattr(item.order, 'created_user', 'N/A'),  # 使用 RMOrder 中的 created_user 或者其他字段作为操作员
        'note': getattr(item.order, 'so_num', '')
    } for item in outbound_actual_logs]

    outbound_stock_logs = OrderItem.objects.filter(product=product,order__is_allocated_to_stock=True).select_related('order')
    outbound_stock_list = [{
        'date': item.order.pickup_date if item.order and item.order.pickup_date else 'N/A',
        'date_shipped': item.order.outbound_date if item.order and item.order.outbound_date else 'N/A',
        'quantity': item.quantity,
        'operator': getattr(item.order, 'created_user', 'N/A'),  # 使用 RMOrder 中的 created_user 或者其他字段作为操作员
        'note': getattr(item.order, 'so_num', '')
    } for item in outbound_stock_logs]

    return inbound_list, outbound_list, outbound_actual_list,outbound_stock_list,inbound_actual_list

# 排序逻辑：将'N/A'视为最小（或最大）值，按需要可修改
def sort_by_date(entry, field_name):
    date_val = entry.get(field_name)
    if date_val is None:
        return datetime.min  # 或 datetime.max，如果想将空值排在最后
    if isinstance(date_val, datetime):
        return date_val
    elif isinstance(date_val, date):
        return datetime.combine(date_val, time.min)
    return datetime.min

@require_http_methods(["POST"])
def order_images(request, order_id):
    try:
        order = RMOrder.objects.get(id=order_id)
        images = request.FILES.getlist('images')  # 获取上传的图片

        for image in images:
            OrderImage.objects.create(order=order, image=image)  # 保存图片与订单的关联

        return JsonResponse({"success": True})
    except RMOrder.DoesNotExist:
        return JsonResponse({"success": False, "error": "Order does not exist."}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)
    
def export_pallet(request):
    # 获取请求中的月份和年份
    month = request.GET.get('month')
    year = request.GET.get('year')

    if month and year:
        # 过滤出指定月份的订单
        start_date = datetime(int(year), int(month), 1)
        if month == '12':
            end_date = datetime(int(year) + 1, 1, 1)  # 下一年1月1日
        else:
            end_date = datetime(int(year), int(month) + 1, 1)  # 下个月的1日

        # 查询 "gloves in" 数据
        gloves_in_orders = Container.objects.filter(empty_date__gte=start_date, empty_date__lt=end_date)

        # 创建 "gloves in" DataFrame
        gloves_in_data = {
            'Empty Date': [order.empty_date for order in gloves_in_orders],  # 假设有 empty_date 字段
            'Container ID': [order.container_id for order in gloves_in_orders],  # 假设有 container_id 字段
            'PLTS': [order.plts for order in gloves_in_orders],
        }
        gloves_in_df = pd.DataFrame(gloves_in_data)

        # 查询 "gloves out" 数据（根据您的需求进行调整）
        gloves_out_orders = RMOrder.objects.filter(outbound_date__gte=start_date, outbound_date__lt=end_date)

        # 创建 "gloves out" DataFrame
        gloves_out_data = {
            'Outbound Date': [order.outbound_date for order in gloves_out_orders],
            'SO': [order.so_num for order in gloves_out_orders],
            'PLTS': [order.plts for order in gloves_out_orders],
        }
        gloves_out_df = pd.DataFrame(gloves_out_data)

        # 创建 Excel 文件
        with pd.ExcelWriter(f'orders_{year}_{month}_pallets.xlsx', engine='openpyxl') as writer:
            gloves_in_df.to_excel(writer, sheet_name='gloves in', index=False)
            gloves_out_df.to_excel(writer, sheet_name='gloves out', index=False)

        # 返回 Excel 文件
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="rmorders_{year}_{month}_pallets.xlsx"'
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            gloves_in_df.to_excel(writer, sheet_name='gloves in', index=False)
            gloves_out_df.to_excel(writer, sheet_name='gloves out', index=False)

        return response
    else:
        return HttpResponse("Invalid month or year", status=400)
    
def import_inventory(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]

        # Read the Excel file into a DataFrame
        df = pd.read_excel(excel_file, engine='openpyxl')

        # Ensure column names match the model fields
        for index, row in df.iterrows():
            # Create RMProduct instance
            product = RMProduct.objects.create(
                name=row["Display Name"],
                shortname = row["Short Name"],
                size = row["Size"],
                description=""  # description 为空
            )
            # Create RMInventory instance
            RMInventory.objects.create(
                product=product,
                quantity=row["Quantity On Hand"],
                quantity_for_neworder=row["Quantity On Hand"],
                quantity_to_stock=row["Quantity On Hand"],
            )

        return JsonResponse({"message": "Excel data imported successfully!"})
    
    return JsonResponse({"error": "No file uploaded"}, status=400)

def import_aline(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]

        # Read only the "Matched Records" sheet
        try:
            df = pd.read_excel(excel_file, sheet_name="Matched Records", engine="openpyxl")
        except ValueError:
            return JsonResponse({"error": "Sheet 'Matched Records' not found in the uploaded file"}, status=400)

        # Loop through each row and save to database
        for _, row in df.iterrows():
            try:
                AlineOrderRecord.objects.create(
                    document_number=row.get("Document Number", ""),  # 文档编号
                    order_number=row.get("Order No", ""),  # 订单编号
                    po_number=row.get("P.O. No.", ""),  # PO编号
                    invoice_date=pd.to_datetime(row.get("Date", None), errors="coerce").date()
                        if pd.notna(row.get("Date", None)) else None,  # 发票日期
                    due_date=pd.to_datetime(row.get("Due Date", None), errors="coerce").date()
                        if pd.notna(row.get("Due Date", None)) else None,  # 截止日期
                    pdf_name=row.get("PDFFileName", ""),  # 文档名称
                    price=row.get("Total", 0) if pd.notna(row.get("Total", 0)) else 0  # 价格
                )
            except Exception as e:
                return JsonResponse({"error": f"Failed to import row {row.to_dict()} - {str(e)}"}, status=400)

        return JsonResponse({"message": "Excel data imported successfully!"})
    
    return JsonResponse({"error": "No file uploaded"}, status=400)

def preview_email(request, number, so_number=None, po_number=None, container_id=None, officedepot_id=None):
    template = "container/temporary.html"
    recipient = "omarorders@omarllc.com,omarwarehouse@rimeius.com"
    current_date = timezone.now().strftime("%m/%d/%Y")
    if number == 1:
        context = {
            "recipient": recipient,
            "subject": f"INVENTORY {current_date}",
            "body": f"Hello,\n\nINVENTORY SUMMARY {current_date}. Paperwork attached.\n\nJing"
        }
    elif number == 2:
        context = {
            "recipient": recipient,
            "subject": "Shipped out Email",
            "body": f"Hello,\n\nSO #{so_number} PO #{po_number}  has been shipped out. Paperwork is attached.\n\nThank you!\nJing"
        }
    elif number == 3:
        context = {
            "recipient": recipient,
            "subject": f"{container_id} RECEIVED IN",
            "body": f"Hello,\n\n{container_id}  has received in. Paperwork is attached.\n\nThank you!\nJing"
        }
    elif number == 4:
        context = {
            "recipient": recipient,
            "subject": f"OFFICE DEPOT #{officedepot_id}",
            "body": f"Hello,\n\nOffice Depot order #{officedepot_id} shipped out. Paperwork is attached.\n\nThank you!\nJing"
        }
    else:
        context = {
            "recipient": recipient,
            "subject": "Received Order Email",
            "body": "Well Received.\n\nThank you!\nJing"
        }
    return render(request, template, context)

def get_user_permissions(user):
    # Use permissionIndex__name to get the name of the permission related to the UserAndPermission instance
    permissions = UserAndPermission.objects.filter(username=user).values_list('permissionIndex__name', flat=True)
    
    # Print the length of the permissions list (or log it)
    print("permissions: ", len(permissions))
    
    return permissions

def export_inventory_to_excel(items):
    wb = openpyxl.Workbook()
    ws = wb.active
    today_str = datetime.now().strftime("%Y-%m-%d")
    ws.title = f"Inventory_{today_str}"

    # Define headers
    headers = ["Product", "Quantity", "New Order", "In Preparation", "Diff","Show Number","Each","Color", "Location", "Pallets", "Cases","ShelfRecord"]
    ws.append(headers)

    # Write data rows
    for item in items:
        ws.append([
            str(item.product),  # Or item.product.name if it's a ForeignKey
            item.quantity,
            item.quantity_for_neworder,
            item.quantity_to_stock,
            item.quantity_diff,
            item.shownumber,
            item.pallet,
            item.color,
            item.Location,
            item.palletnumber,
            item.case,
            item.ShelfRecord,            
        ])

    # Auto-fit column width
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Save to in-memory file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create HTTP response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Inventory_{today_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
