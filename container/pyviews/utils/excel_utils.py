from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def format_worksheet(ws, colors=None, color_map=None):
    """格式化 Excel 工作表：设置列宽、对齐、边框、颜色"""
    alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        
        for cell in col:
            cell.alignment = alignment
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = max(max_length + 2, 12)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 如果提供了颜色映射，应用颜色
    if color_map and colors:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.row > 1:  # 跳过表头
                    employee_name = ws.cell(row=cell.row, column=1).value
                    employee_color = color_map.get(employee_name)
                    if employee_color:
                        cell.fill = PatternFill(
                            start_color=employee_color,
                            end_color=employee_color,
                            fill_type="solid"
                        )
                    cell.border = thin_border