import os
import pandas as pd

from datetime import date, datetime, timedelta,time
from openpyxl.styles import Border, Side, PatternFill, Alignment

from django.utils import timezone
from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.http import HttpResponse

from ..models import ClockRecord,Employee,UserAndPermission


def week_record(request):
    # 获取当前年份和周数（ISO标准）
    current_year, current_week, _ = date.today().isocalendar()
    print("current_week:", current_year, current_week)

    # 本周的开始（周一）和结束（周日）日期
    current_week_start = date.fromisocalendar(current_year, current_week, 1)
    current_week_end = current_week_start + timedelta(days=6)
    print("Week Range:", current_week_start, current_week_end)

    # 上一周的周数和年份（注意处理跨年情况）
    last_week_date = current_week_start - timedelta(days=7)
    last_week_year, last_week_number, _ = last_week_date.isocalendar()

    # 获取前端传入的 year 和 week，默认值为当前年的上一周
    selected_year = int(request.GET.get("year", last_week_year))
    selected_week = int(request.GET.get("week", last_week_number))

    # 获取选定周的起止日期
    selected_week_start = date.fromisocalendar(selected_year, selected_week, 1)
    selected_week_end = selected_week_start + timedelta(days=6)

    print("Selected Week:", selected_year, selected_week)
    print("Week Range:", selected_week_start, selected_week_end)

    # 获取选定周的打卡记录
    weekly_records = ClockRecord.objects.filter(date__range=[selected_week_start, selected_week_end])

    # 生成年份和周数选项
    years = list(range(current_year, current_year + 1))  # 当前年份及前两年
    weeks = list(range(1, 53))  # 1-52周

    user_permissions = get_user_permissions(request.user)

    # 如果没有记录，返回空页面或消息
    if not weekly_records.exists():
        return render(request, 'container/weekrecord.html', {
            'employee_records': None,  # 或者可以返回一个空的字典
            'years': years,
            'weeks': weeks,
            'selected_year': selected_year,
            'selected_week': selected_week,
            'user_permissions': user_permissions
        })
    
    # 获取所有员工的周统计数据
    employees = Employee.objects.filter(belongTo="CabinetsDepot")
    employee_records = []
    
    for employee in employees:
        employee_weekly_records = weekly_records.filter(employee_name=employee)
        employee_total_hours = sum(record.total_hours for record in employee_weekly_records)
        employee_avg_hours = round(employee_total_hours / 5, 2) if employee_weekly_records else 0
        employee_attendance_rate = round((employee_total_hours / 40) * 100, 2)
        
        # 只有当员工有打卡记录时才添加到列表中
        if employee_weekly_records.exists():
            employee_records.append({
                'id':employee.id,
                'name': employee.name,
                'period_start': selected_week_start,
                'period_end': selected_week_end,
                'total_hours': employee_total_hours,
                'average_hours': employee_avg_hours,
                "attendance_rate":employee_attendance_rate,
                'belongTo':employee.belongTo
            })
        else:
            employee_records.append({
                'id':employee.id,
                'name': employee.name,
                'period_start': selected_week_start,
                'period_end': selected_week_end,
                'total_hours': 0,
                'average_hours': 0,
                "attendance_rate":0,
                'belongTo':employee.belongTo
            })

    # ✅ 排序
    employee_records = sorted(employee_records, key=lambda x: (x['belongTo'] or '').lower())
    
    return render(request, 'container/weekrecord.html', {
        'employee_records': employee_records,
        'years': years,
        'weeks': weeks,
        'selected_year': selected_year,
        'selected_week': selected_week,
        'user_permissions': user_permissions
    })

def add_week_records(request):
    today = timezone.now().date()
    current_week_start = today - timedelta(days=today.weekday())
    # 获取上周的周一日期
    last_week_start = current_week_start - timedelta(days=7)  # 上周的周一
    last_week_end = last_week_start + timedelta(days=6)  # 上周的周日
    print("hello: ",last_week_start,last_week_end)
    
    weekdays = getWeek(last_week_start)
    
    if request.method == 'POST':
        # 获取选择的员工名称
        name = request.POST.get('employee_name')
        print(f"Selected Employee Name: {name}")
        employee = Employee.objects.get(id=name)
        print(f"Selected Employee Name: {employee.name}")

        # 创建新的打卡记录
        # clock_record = ClockRecord(
        #     employee_name=employee,
        #     date=date.today(),  # 今天的日期
        #     weekday=date.today().weekday(),  # 自动获取星期几（0-6）
        #     morning_in="09:00",  # 可以直接使用字符串，save方法会自动转换
        #     morning_out="12:00",
        #     afternoon_in="13:00",
        #     afternoon_out="18:00"
        #     # evening_in 和 evening_out 是可选的，可以不填
        # )

        # # 保存到数据库
        # clock_record.save()
        
        for day in weekdays:
            # weekday = "Monday"
            weekdayname = day['name']
            weekday = day['weekday']
            date = day['date']            
            
            # 获取表单数据
            morning_in = request.POST.get(f'morning_in_{weekday}')
            morning_out = request.POST.get(f'morning_out_{weekday}')
            afternoon_in = request.POST.get(f'afternoon_in_{weekday}')
            afternoon_out = request.POST.get(f'afternoon_out_{weekday}')
            evening_in = request.POST.get(f'evening_in_{weekday}')
            evening_out = request.POST.get(f'evening_out_{weekday}')
            print("I am ok now\n", employee.name,date, weekdayname)
            print("Time: ",morning_in, morning_out, afternoon_in, afternoon_out, evening_in, evening_out )
            
            # 创建记录
            if any([morning_in, morning_out, afternoon_in, afternoon_out, evening_in, evening_out]):
                ClockRecord.objects.create(
                    date=date,
                    weekday=weekday,
                    employee_name=employee,
                    morning_in=morning_in or None,
                    morning_out=morning_out or None,
                    afternoon_in=afternoon_in or None,
                    afternoon_out=afternoon_out or None,
                    evening_in=evening_in or None,
                    evening_out=evening_out or None
                )
                # print("hello: ",type(wordrecord))
                # wordrecord.save()
                # return JsonResponse({"status": "success", "message": "Record added!"})
        
        return redirect('week_record')
    
    employees = Employee.objects.all()
    return render(request, 'container/weekrecord/add_record.html', {
        'weekdays': weekdays,
        'employees': employees,
    })

def edit_week_records(request, employee_id=None):
    today = timezone.now().date()
    current_week_start = today - timedelta(days=today.weekday())
    # 获取上周的周一日期
    last_week_start = current_week_start - timedelta(days=7)
    employee = Employee.objects.get(id=employee_id)  
    
    weekdays = getWeek(last_week_start)

    if request.method == 'POST':
        # Handle form submission to save work records
        for i in range(7):
            morning_in = request.POST.get(f'morning_in_{i}')
            morning_out = request.POST.get(f'morning_out_{i}')
            afternoon_in = request.POST.get(f'afternoon_in_{i}')
            afternoon_out = request.POST.get(f'afternoon_out_{i}')
            evening_in = request.POST.get(f'evening_in_{i}')
            evening_out = request.POST.get(f'evening_out_{i}')

            # Get weekday value from weekdays list
            weekday = weekdays[i]['weekday']
            date = last_week_start + timedelta(days=i)  # Calculate the date for the current weekday    
            print("hello: ",employee.name, weekdays[i]['name'], date)
            print("hello: ",morning_in,morning_out,afternoon_in,afternoon_out,evening_in,evening_out)
            
            # Save or update work records
            ClockRecord.objects.update_or_create(
                employee_name = employee,
                date = date,
                defaults={
                    'weekday': weekday,
                    'morning_in': morning_in or None,
                    'morning_out': morning_out or None,
                    'afternoon_in': afternoon_in or None,
                    'afternoon_out': afternoon_out or None,
                    'evening_in': evening_in or None,
                    'evening_out': evening_out or None,
                }
            )
            # 先检查是否存在记录
            # record, created = ClockRecord.objects.get_or_create(
            #     employee_name=employee,
            #     date=date,
            # )
            
            # # 直接更新字段
            # record.weekday = weekday
            # record.morning_in = morning_in if morning_in else record.morning_in
            # record.morning_out = morning_out if morning_out else record.morning_out
            # record.afternoon_in = afternoon_in if afternoon_in else record.afternoon_in
            # record.afternoon_out = afternoon_out if afternoon_out else record.afternoon_out
            # record.evening_in = evening_in if evening_in else record.evening_in
            # record.evening_out = evening_out if evening_out else record.evening_out
            # record.save()
        return redirect('week_record')

    current_employee = None
    work_records = []
    if employee_id:
        current_employee = Employee.objects.get(id=employee_id)
        work_records = ClockRecord.objects.filter(employee_name=current_employee, date__range=[last_week_start, last_week_start + timedelta(days=6)])
    
    work_record_dict = {record.date: record for record in work_records}

    return render(request, 'container/weekrecord/edit_record.html', {
        'weekdays': weekdays,
        'current_employee': current_employee,
        'work_records': work_record_dict,  # Pass the dictionary to the template
    })

def export_week_records(request):

    # 获取所有员工
    employees = Employee.objects.all()

    # 获取前端传来的 group 参数，例如 Aline 或 CabinetsDepot
    group_param = request.GET.get('brand')
    year = int(request.GET.get("year"))
    week = int(request.GET.get("week"))
    select_week_start = date.fromisocalendar(year, week, 1)  # 周一
    select_week_end =  date.fromisocalendar(year, week, 7)    # 周日
    weekdays = getWeek(select_week_start)
    print("export_week_records:",group_param)
    print("year: ",year," week: ",week)

    # 按照 belongTo 字段分组员工
    grouped_employees = {}
    for employee in employees:
        group = employee.belongTo  # 假设 belongTo 是一个字符串，表示员工所属的组
        if group not in grouped_employees:
            grouped_employees[group] = []
        grouped_employees[group].append(employee)    
    
    if group_param:
        # 只保留匹配组的员工
        grouped_employees = {group_param: grouped_employees.get(group_param, [])}

    # 定义颜色列表
    colors = ["FFFF99", "FFCCFF", "CCFFCC", "FFCCCC", "CCCCFF"]
    color_map = {}  # 用于存储每个员工的颜色

    # 为每个员工分配颜色
    for index, employee in enumerate(employees):
        color_map[employee.name] = colors[index % len(colors)]  # 循环使用颜色

    # 存储所有响应
    responses = []

    # 为每个组创建一个工作表
    for group, group_employees in grouped_employees.items():
        print("---group: ",group)
        filename = f'Working_Hours_{group}_{select_week_start.strftime("%m.%d")}-{select_week_end.strftime("%m.%d")}.2025.xlsx'
        # 拼接完整路径
        save_dir = os.path.join(os.getcwd(), 'media/workrecord')
        os.makedirs(save_dir, exist_ok=True)
        full_path = os.path.join(save_dir, filename)
        
        data = []        
        with pd.ExcelWriter(full_path, engine='openpyxl') as writer: 
            for employee in group_employees:
                total_hours_weekly = 0
                employee_records = []

                # 获取该员工的上周打卡记录
                records = ClockRecord.objects.filter(employee_name=employee, date__range=[select_week_start, select_week_end]).order_by('date')

                for record in records:
                    # 计算每天的工作时长
                    morning_in = record.morning_in
                    morning_out = record.morning_out
                    afternoon_in = record.afternoon_in
                    afternoon_out = record.afternoon_out
                    evening_in = record.evening_in
                    evening_out = record.evening_out
                    total_hours = record.total_hours
                    total_hours_weekly += total_hours

                    # 将 weekday 数字转换为字符串
                    weekday_str = weekdays[record.weekday]['name']

                    employee_records.append({
                    'Name': employee.name,
                    'Date': record.date,
                    'Weekday': weekday_str,
                    'In Time1': morning_in.strftime('%H:%M') if morning_in else '',
                    'Out Time1': morning_out.strftime('%H:%M') if morning_out else '',
                    'In Time2': afternoon_in.strftime('%H:%M') if afternoon_in else '',
                    'Out Time2': afternoon_out.strftime('%H:%M') if afternoon_out else '',
                    'In Time3': evening_in.strftime('%H:%M') if evening_in else '',
                    'Out Time3': evening_out.strftime('%H:%M') if evening_out else '',
                    'Total Hours Daily': total_hours,
                    'Total Hours Weekly': ''
                    })  

                # 在最后一条记录中添加周总时间
                if employee_records:
                    employee_records[-1]['Total Hours Weekly'] = total_hours_weekly 
                    data.extend(employee_records)                  
            if data:
                # 创建 DataFrame
                df = pd.DataFrame(data)

                # 将 DataFrame 写入 Excel 的一个工作表
                df.to_excel(writer, sheet_name=group, index=False)

                # 获取当前工作表
                worksheet = writer.sheets[group]

                # 设置列宽以适应内容
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                # 设置不同颜色和边框
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=len(data[0])):
                    for cell in row:
                        # 使用当前行员工的颜色
                        employee_id = data[row[0].row - 2]['Name']  # 获取当前行的员工名称
                        employee_color = color_map.get(employee_id)  # 获取该员工的颜色
                        # print("color2: ",employee_id, employee_color)
                        cell.fill = PatternFill(start_color=employee_color, end_color=employee_color, fill_type="solid")
                        cell.border = thin_border  # 添加边框
                        cell.alignment = Alignment(horizontal='center', vertical='center')  # 设置居中对齐

        # 返回 Excel 文件作为响应
        if os.path.exists(full_path):
            with open(full_path, 'rb') as f:
                file_data = f.read()
                response = HttpResponse( 
                    file_data,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
    
        # 如果没有数据，返回空响应
        return HttpResponse("No data available for the selected group.")

def convertToTime(workTime):
    if isinstance(workTime, str):
        print("Convert string to time")
        workTime = datetime.strptime(workTime, '%H:%M').time()  # Convert string to time object

    elif isinstance(workTime, datetime):
        print("Extract time from datetime")
        workTime = workTime.time()  # Extract only time from datetime object

    elif not isinstance(workTime, time):
        print("Set default time (08:00)")
        workTime = time(8, 0)  # Default to 8:00 AM if invalid

    return workTime

def getWeek(last_week_start):
    weekdays = []
    for i in range(7):  # Get all days of the week
        weekday_date = last_week_start + timedelta(days=i)
        weekdays.append({
            'weekday': i,
            'name': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'][i],
            'date': weekday_date
        })
    return weekdays

def get_user_permissions(user):
    # Use permissionIndex__name to get the name of the permission related to the UserAndPermission instance
    permissions = UserAndPermission.objects.filter(username=user).values_list('permissionIndex__name', flat=True)
    
    # Print the length of the permissions list (or log it)
    print("permissions: ", len(permissions))
    
    return permissions